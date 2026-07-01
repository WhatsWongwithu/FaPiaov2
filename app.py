"""
发票识别系统 - Web客户端
上海辉驰包装设备有限公司 财务部

启动后浏览器访问 http://localhost:5000
"""

import os
import base64
import time
import uuid
import requests
from flask import Flask, request, jsonify, send_file, render_template
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from config import get_baidu_keys

# ==================== 配置 ====================
API_KEY, SECRET_KEY = get_baidu_keys()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
PDF_EXTENSIONS = (".pdf",)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB


# ==================== 百度OCR ====================
def get_access_token():
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {
        "grant_type": "client_credentials",
        "client_id": API_KEY,
        "client_secret": SECRET_KEY,
    }
    resp = requests.post(url, params=params, timeout=10)
    resp.raise_for_status()
    return resp.json()["access_token"]


_token = None
_token_time = 0


def get_token():
    global _token, _token_time
    if not _token or time.time() - _token_time > 86400:
        _token = get_access_token()
        _token_time = time.time()
    return _token


def ocr_vat_invoice(file_path):
    token = get_token()
    url = (
        "https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice"
        f"?access_token={token}"
    )
    with open(file_path, "rb") as f:
        file_b64 = base64.b64encode(f.read()).decode("utf-8")

    ext = os.path.splitext(file_path)[1].lower()
    if ext in PDF_EXTENSIONS:
        data = {"pdf_file": file_b64, "pdf_file_num": "1"}
    else:
        data = {"image": file_b64}

    resp = requests.post(url, data=data, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ==================== 解析 ====================
def extract_value(val):
    if isinstance(val, dict):
        return val.get("word", val.get("words", ""))
    if isinstance(val, list):
        return [extract_value(v) for v in val]
    return val if val else ""


def safe_get(obj, key, default=""):
    if isinstance(obj, dict):
        return extract_value(obj.get(key, default))
    return default


def parse_invoice(result):
    wr = result.get("words_result", {})

    date = safe_get(wr, "InvoiceDate")
    invoice_num = safe_get(wr, "InvoiceNum")
    total_amount = safe_get(wr, "TotalAmount")
    seller_name = safe_get(wr, "SellerName")

    names = safe_get(wr, "CommodityName", [])
    specs = safe_get(wr, "CommoditySpec", [])
    nums = safe_get(wr, "CommodityNum", [])
    prices = safe_get(wr, "CommodityPrice", [])

    if isinstance(names, str):
        names = [names]
    if isinstance(specs, str):
        specs = [specs]
    if isinstance(nums, str):
        nums = [nums]
    if isinstance(prices, str):
        prices = [prices]

    n = max(len(names), len(specs), len(nums), len(prices), 1)

    items = []
    for i in range(n):
        items.append(
            {
                "name": names[i] if i < len(names) else "",
                "spec": specs[i] if i < len(specs) else "",
                "qty": nums[i] if i < len(nums) else "",
                "price": prices[i] if i < len(prices) else "",
            }
        )

    return {
        "date": date,
        "invoice_num": invoice_num,
        "total_amount": total_amount,
        "seller_name": seller_name,
        "items": items,
    }


# ==================== Excel生成 ====================
def write_excel(invoices, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "进项发票记录"

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="微软雅黑", bold=True, size=10)
    title_font = Font(name="微软雅黑", bold=True, size=14)
    data_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill(
        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:M1")
    ws["A1"] = "进项发票记录及应付账款"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    headers = [
        "序号", "开票日期", "品名", "规格", "数量", "单价",
        "发票总金额", "票号", "已付款", "未付款", "付款方式",
        "付款日期", "供应商名称",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    row = 3
    serial = 1
    for inv in invoices:
        first_row = True
        for item in inv["items"]:
            if first_row:
                ws.cell(row=row, column=1, value=serial)
                ws.cell(row=row, column=2, value=inv["date"])
                ws.cell(row=row, column=7, value=inv["total_amount"])
                ws.cell(row=row, column=8, value=f"NO.{inv['invoice_num']}")
                ws.cell(row=row, column=10, value=inv["total_amount"])
                ws.cell(row=row, column=13, value=inv["seller_name"])
                first_row = False
            else:
                ws.cell(row=row, column=10, value=0)
                ws.cell(row=row, column=13, value=inv["seller_name"])

            ws.cell(row=row, column=3, value=item["name"])
            ws.cell(row=row, column=4, value=item["spec"])
            ws.cell(row=row, column=5, value=item["qty"])
            ws.cell(row=row, column=6, value=item["price"])
            row += 1
        serial += 1

    max_row = row - 1
    for r in range(3, max_row + 1):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = left_align if c in (3, 4) else center
            cell.font = data_font

    col_widths = [6, 12, 14, 18, 8, 10, 14, 26, 10, 10, 10, 12, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    sum_row = max_row + 1
    ws.cell(row=sum_row, column=6, value="合计")
    ws.cell(row=sum_row, column=6).font = header_font
    ws.cell(row=sum_row, column=6).alignment = center
    ws.cell(row=sum_row, column=7, value=f"=SUM(G3:G{max_row})")
    ws.cell(row=sum_row, column=7).font = header_font
    ws.cell(row=sum_row, column=7).alignment = center
    for c in range(1, 14):
        ws.cell(row=sum_row, column=c).border = thin_border

    wb.save(output_path)


# ==================== 路由 ====================
# 存储已识别的发票（内存中，服务重启后清空）
invoices_store = []


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
def upload():
    """上传文件并识别"""
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未收到文件"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in IMAGE_EXTENSIONS and ext not in PDF_EXTENSIONS:
        return jsonify({"error": f"不支持的格式: {ext}"}), 400

    # 保存文件
    safe_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    file.save(file_path)

    # OCR识别
    try:
        result = ocr_vat_invoice(file_path)
        if result.get("error_code"):
            return jsonify(
                {"error": f"识别失败: {result.get('error_msg', '未知错误')}"}
            ), 200

        if "words_result" not in result:
            return jsonify({"error": "未识别到发票内容"}), 200

        invoice = parse_invoice(result)
        invoice["id"] = uuid.uuid4().hex
        invoice["filename"] = file.filename
        invoices_store.append(invoice)

        return jsonify(
            {
                "success": True,
                "invoice": {
                    "id": invoice["id"],
                    "filename": invoice["filename"],
                    "date": invoice["date"],
                    "invoice_num": invoice["invoice_num"],
                    "total_amount": invoice["total_amount"],
                    "seller_name": invoice["seller_name"],
                    "item_count": len(invoice["items"]),
                    "items": invoice["items"],
                },
            }
        )
    except Exception as e:
        return jsonify({"error": f"处理异常: {str(e)}"}), 200
    finally:
        # 清理临时文件
        if os.path.exists(file_path):
            os.remove(file_path)


@app.route("/results")
def results():
    """获取所有已识别的发票"""
    return jsonify(
        {
            "invoices": [
                {
                    "id": inv["id"],
                    "filename": inv["filename"],
                    "date": inv["date"],
                    "invoice_num": inv["invoice_num"],
                    "total_amount": inv["total_amount"],
                    "seller_name": inv["seller_name"],
                    "item_count": len(inv["items"]),
                    "items": inv["items"],
                }
                for inv in invoices_store
            ]
        }
    )


@app.route("/download")
def download():
    """生成并下载Excel"""
    if not invoices_store:
        return jsonify({"error": "没有数据"}), 400

    filename = f"进项发票记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(UPLOAD_DIR, filename)
    write_excel(invoices_store, output_path)

    return send_file(
        output_path,
        as_attachment=True,
        download_name=filename,
    )


@app.route("/clear")
def clear():
    """清空所有已识别的发票"""
    invoices_store.clear()
    return jsonify({"success": True})


@app.route("/delete/<invoice_id>")
def delete_invoice(invoice_id):
    """删除单条发票"""
    global invoices_store
    invoices_store = [inv for inv in invoices_store if inv["id"] != invoice_id]
    return jsonify({"success": True})


if __name__ == "__main__":
    print("=" * 50)
    print("  发票识别系统 Web客户端")
    print("  上海辉驰包装设备有限公司 财务部")
    print("=" * 50)
    print("\n  浏览器访问: http://localhost:8080\n")
    app.run(host="0.0.0.0", port=8080, debug=True)
