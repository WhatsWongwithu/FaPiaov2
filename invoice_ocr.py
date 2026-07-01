"""
发票拍照识别与自动归类系统
上海辉驰包装设备有限公司 财务部

使用方法：
  python3 invoice_ocr.py

将发票照片放入「发票照片」文件夹，运行后自动生成「进项发票记录.xlsx」
"""

import os
import sys
import base64
import time
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from config import get_baidu_keys

# ==================== 配置 ====================
API_KEY, SECRET_KEY = get_baidu_keys()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PHOTO_DIR = os.path.join(BASE_DIR, "发票照片")
OUTPUT_FILE = os.path.join(BASE_DIR, f"进项发票记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
PDF_EXTENSIONS = (".pdf",)
ALL_EXTENSIONS = IMAGE_EXTENSIONS + PDF_EXTENSIONS


# ==================== 百度OCR ====================
def get_access_token():
    """获取百度OCR access_token"""
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {
        "grant_type": "client_credentials",
        "client_id": API_KEY,
        "client_secret": SECRET_KEY,
    }
    resp = requests.post(url, params=params, timeout=10)
    resp.raise_for_status()
    return resp.json()["access_token"]


def ocr_vat_invoice(file_path, access_token):
    """调用增值税发票识别接口（支持图片和PDF）"""
    url = (
        "https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice"
        f"?access_token={access_token}"
    )

    with open(file_path, "rb") as f:
        file_b64 = base64.b64encode(f.read()).decode("utf-8")

    ext = os.path.splitext(file_path)[1].lower()
    if ext in PDF_EXTENSIONS:
        # PDF文件：使用pdf_file参数，识别第1页
        data = {"pdf_file": file_b64, "pdf_file_num": "1"}
    else:
        # 图片文件：使用image参数
        data = {"image": file_b64}

    resp = requests.post(url, data=data, timeout=30)
    resp.raise_for_status()
    return resp.json()


# ==================== 解析结果 ====================
def extract_value(val):
    """从百度返回值中提取纯文本（处理str/dict/list混合格式）"""
    if isinstance(val, dict):
        return val.get("word", val.get("words", ""))
    if isinstance(val, list):
        return [extract_value(v) for v in val]
    return val if val else ""


def safe_get(obj, key, default=""):
    """安全获取字段值"""
    if isinstance(obj, dict):
        val = obj.get(key, default)
        return extract_value(val)
    return default


def parse_invoice(result):
    """将百度OCR返回结果解析为统一结构"""
    wr = result.get("words_result", {})

    # 发票基本字段
    date = safe_get(wr, "InvoiceDate")
    invoice_num = safe_get(wr, "InvoiceNum")
    total_amount = safe_get(wr, "TotalAmount")  # 价税合计
    seller_name = safe_get(wr, "SellerName")

    # 明细行字段（百度返回数组）
    names = safe_get(wr, "CommodityName", [])
    specs = safe_get(wr, "CommoditySpec", [])
    nums = safe_get(wr, "CommodityNum", [])
    prices = safe_get(wr, "CommodityPrice", [])

    if isinstance(names, str):
        names = [names]
    if isinstance(specs, str):
        specs = [specs]
    if isinstance(nums, str):
        nums = [nums]
    if isinstance(prices, str):
        prices = [prices]

    # 确定明细行数
    n = max(len(names), len(specs), len(nums), len(prices), 1)

    items = []
    for i in range(n):
        items.append(
            {
                "name": names[i] if i < len(names) else "",
                "spec": specs[i] if i < len(specs) else "",
                "qty": nums[i] if i < len(nums) else "",
                "price": prices[i] if i < len(prices) else "",
            }
        )

    return {
        "date": date,
        "invoice_num": invoice_num,
        "total_amount": total_amount,
        "seller_name": seller_name,
        "items": items,
    }


# ==================== 写入Excel ====================
def write_excel(invoices, output_path):
    """按模板格式写入Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "进项发票记录"

    # ---- 样式 ----
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="微软雅黑", bold=True, size=10)
    title_font = Font(name="微软雅黑", bold=True, size=14)
    data_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---- 第1行：标题 ----
    ws.merge_cells("A1:M1")
    ws["A1"] = "进项发票记录及应付账款"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    # ---- 第2行：表头 ----
    headers = [
        "序号", "开票日期", "品名", "规格", "数量", "单价",
        "发票总金额", "票号", "已付款", "未付款", "付款方式",
        "付款日期", "供应商名称",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    # ---- 写入数据 ----
    row = 3
    serial = 1
    for inv in invoices:
        first_row = True
        for item in inv["items"]:
            if first_row:
                ws.cell(row=row, column=1, value=serial)
                ws.cell(row=row, column=2, value=inv["date"])
                ws.cell(row=row, column=7, value=inv["total_amount"])
                ws.cell(row=row, column=8, value=f"NO.{inv['invoice_num']}")
                ws.cell(row=row, column=10, value=inv["total_amount"])
                ws.cell(row=row, column=13, value=inv["seller_name"])
                first_row = False
            else:
                ws.cell(row=row, column=10, value=0)
                ws.cell(row=row, column=13, value=inv["seller_name"])

            ws.cell(row=row, column=3, value=item["name"])
            ws.cell(row=row, column=4, value=item["spec"])
            ws.cell(row=row, column=5, value=item["qty"])
            ws.cell(row=row, column=6, value=item["price"])

            row += 1
        serial += 1

    # ---- 统一格式 ----
    max_row = row - 1
    for r in range(3, max_row + 1):
        for c in range(1, 14):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            if c in (3, 4):  # 品名、规格
                cell.alignment = left_align
            else:
                cell.alignment = center
            if not cell.font or cell.font.name != "微软雅黑":
                cell.font = data_font

    # ---- 列宽 ----
    col_widths = [6, 12, 14, 18, 8, 10, 14, 26, 10, 10, 10, 12, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ---- 合计行 ----
    sum_row = max_row + 1
    ws.cell(row=sum_row, column=6, value="合计")
    ws.cell(row=sum_row, column=6).font = header_font
    ws.cell(row=sum_row, column=6).alignment = center
    ws.cell(row=sum_row, column=7, value=f"=SUM(G3:G{max_row})")
    ws.cell(row=sum_row, column=7).font = header_font
    ws.cell(row=sum_row, column=7).alignment = center
    for c in range(1, 14):
        ws.cell(row=sum_row, column=c).border = thin_border

    wb.save(output_path)


# ==================== 主流程 ====================
def main():
    print("=" * 50)
    print("  发票拍照识别与自动归类系统")
    print("  上海辉驰包装设备有限公司 财务部")
    print("=" * 50)

    # 检查照片文件夹
    if not os.path.exists(PHOTO_DIR):
        os.makedirs(PHOTO_DIR)
        print(f"\n已创建文件夹: {PHOTO_DIR}")
        print("请将发票照片放入该文件夹后重新运行。")
        input("按回车键退出...")
        return

    # 收集图片和PDF
    files = [
        os.path.join(PHOTO_DIR, f)
        for f in sorted(os.listdir(PHOTO_DIR))
        if f.lower().endswith(ALL_EXTENSIONS)
    ]

    if not files:
        print(f"\n未在「发票照片」文件夹中找到图片或PDF文件。")
        print(f"请将发票照片(.jpg/.png)或PDF放入: {PHOTO_DIR}")
        input("按回车键退出...")
        return

    print(f"\n找到 {len(files)} 个文件，开始识别...\n")

    # 获取access_token
    print("正在连接百度OCR服务...", end=" ")
    try:
        token = get_access_token()
        print("✅ 连接成功")
    except Exception as e:
        print(f"❌ 连接失败: {e}")
        input("按回车键退出...")
        return

    # 逐张识别
    invoices = []
    success = 0
    fail = 0

    for i, img_path in enumerate(files, 1):
        fname = os.path.basename(img_path)
        print(f"[{i}/{len(files)}] 识别中: {fname}", end=" ")

        try:
            result = ocr_vat_invoice(img_path, token)

            if result.get("error_code"):
                err_msg = result.get("error_msg", "未知错误")
                print(f"❌ 失败 ({err_msg})")
                fail += 1
                continue

            if "words_result" not in result:
                print("❌ 未识别到发票内容")
                fail += 1
                continue

            invoice = parse_invoice(result)
            invoices.append(invoice)
            print(f"✅ {invoice['seller_name']} | {invoice['invoice_num']} | {len(invoice['items'])}条明细")
            success += 1

            # 百度API限速：QPS=2，适当间隔
            if i < len(files):
                time.sleep(0.6)

        except Exception as e:
            print(f"❌ 异常: {e}")
            fail += 1

    # 写入Excel
    if invoices:
        print(f"\n识别完成: 成功 {success} 张, 失败 {fail} 张")
        print("正在生成Excel...", end=" ")
        write_excel(invoices, OUTPUT_FILE)
        print("✅")
        print(f"\n📁 文件已保存: {OUTPUT_FILE}")
        print(f"   共 {len(invoices)} 张发票, {sum(len(i['items']) for i in invoices)} 条明细")
    else:
        print(f"\n未能识别任何发票，请检查照片清晰度后重试。")

    print("\n" + "=" * 50)
    input("按回车键退出...")


if __name__ == "__main__":
    main()
