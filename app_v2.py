"""
发票识别系统 V3 - OCR + DeepSeek LLM解析
上海辉驰包装设备有限公司 财务部

OCR识别文字 → DeepSeek理解表格结构 → 提取全部字段
启动后浏览器访问 http://localhost:8080
"""

import os
import json
import uuid
import sqlite3
from functools import wraps
from flask import (
    Flask, request, jsonify, send_file, render_template,
    session, redirect,
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta

from ocr_engine import OCREngine
from llm_parser import parse_with_deepseek
from config import get_deepseek_key, get_accounts

FALLBACK_API_KEY = get_deepseek_key()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
DB_PATH = os.path.join(BASE_DIR, "history.db")

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff")
PDF_EXTENSIONS = (".pdf",)
ALL_EXTENSIONS = IMAGE_EXTENSIONS + PDF_EXTENSIONS

app = Flask(__name__)
app.secret_key = "huichi-fapiao-2026-secret-key-fixed"
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.permanent_session_lifetime = timedelta(hours=24)

ocr_engine = OCREngine()


# ==================== 数据库 ====================
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id            TEXT PRIMARY KEY,
            invoice_num   TEXT,
            date          TEXT,
            seller_name   TEXT,
            total_amount  TEXT,
            items         TEXT,
            filename      TEXT,
            uploaded_by   TEXT,
            created_at    TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id                INTEGER PRIMARY KEY AUTOINCREMENT,
            username          TEXT UNIQUE NOT NULL,
            password_hash     TEXT NOT NULL,
            deepseek_api_key  TEXT DEFAULT '',
            created_at        TEXT
        )
    """)
    # 初始化2个固定账号
    accounts = get_accounts()
    for acc in accounts:
        existing = conn.execute("SELECT 1 FROM users WHERE username = ?", (acc["username"],)).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO users (username, password_hash, deepseek_api_key, created_at) VALUES (?,?,?,?)",
                (acc["username"], generate_password_hash(acc["password"]), "", datetime.now().isoformat()),
            )
            print(f"  账号: {acc['username']} / {acc['password']}")
    conn.commit()
    conn.close()


def get_user_api_key(user_id):
    """从数据库取当前用户的 DeepSeek API Key，没有则返回兜底Key"""
    conn = get_db()
    row = conn.execute("SELECT deepseek_api_key FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.close()
    if row and row["deepseek_api_key"]:
        return row["deepseek_api_key"]
    return FALLBACK_API_KEY


# ==================== 认证装饰器 ====================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            if request.path.startswith("/api/") or request.is_json:
                return jsonify({"error": "未登录", "redirect": "/login"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


# ==================== Excel生成 ====================
def _to_number(val):
    """将字符串转为数字（int或float），失败返回None"""
    if val is None or str(val).strip() == "":
        return None
    try:
        s = str(val).replace(",", "").replace("￥", "").replace("¥", "").strip()
        f = float(s)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return None


def _to_date(val):
    """将日期字符串转为 datetime 对象，供Excel识别为日期类型"""
    if not val:
        return None
    s = str(val).strip().replace("-", "/")
    for fmt in ("%Y/%m/%d", "%Y/%m/%d %H:%M", "%Y年%m月%d日"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return val


def write_excel(invoices, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "进项发票记录"

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_font = Font(name="微软雅黑", bold=True, size=10)
    title_font = Font(name="微软雅黑", bold=True, size=14)
    data_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    num_cols = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws["A1"] = "进项发票记录"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    headers = ["序号", "开票日期", "品名", "规格", "数量", "单价",
               "金额", "税率", "税额", "发票总金额", "票号", "供应商名称"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        cell.fill = header_fill

    row = 3
    serial = 1
    for inv in invoices:
        for item in inv["items"]:
            ws.cell(row=row, column=1, value=serial)
            date_cell = ws.cell(row=row, column=2, value=_to_date(inv["date"]))
            date_cell.number_format = "YYYY/MM/DD"
            ws.cell(row=row, column=3, value=item["name"])
            ws.cell(row=row, column=4, value=item["spec"])
            ws.cell(row=row, column=5, value=_to_number(item["qty"]))
            ws.cell(row=row, column=6, value=_to_number(item["price"]))
            ws.cell(row=row, column=7, value=_to_number(item["amount"]))
            ws.cell(row=row, column=8, value=item["tax_rate"])
            ws.cell(row=row, column=9, value=_to_number(item["tax"]))
            ws.cell(row=row, column=10, value=_to_number(inv["total_amount"]))
            ws.cell(row=row, column=11, value=f"NO.{inv['invoice_num']}")
            ws.cell(row=row, column=12, value=inv["seller_name"])
            row += 1
        serial += 1

    max_row = row - 1
    for r in range(3, max_row + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = left_align if c in (3, 4, 12) else center
            cell.font = data_font

    col_widths = [6, 12, 16, 18, 8, 14, 12, 8, 12, 14, 26, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    sum_row = max_row + 1
    ws.cell(row=sum_row, column=3, value="合计")
    ws.cell(row=sum_row, column=3).font = header_font
    ws.cell(row=sum_row, column=3).alignment = center
    ws.cell(row=sum_row, column=7, value=f"=SUM(G3:G{max_row})")
    ws.cell(row=sum_row, column=7).font = header_font
    ws.cell(row=sum_row, column=7).alignment = center
    ws.cell(row=sum_row, column=9, value=f"=SUM(I3:I{max_row})")
    ws.cell(row=sum_row, column=9).font = header_font
    ws.cell(row=sum_row, column=9).alignment = center
    ws.cell(row=sum_row, column=10, value=f"=SUM(J3:J{max_row})")
    ws.cell(row=sum_row, column=10).font = header_font
    ws.cell(row=sum_row, column=10).alignment = center
    for c in range(1, num_cols + 1):
        ws.cell(row=sum_row, column=c).border = thin_border

    tab = Table(displayName="InvoiceTable", ref=f"A2:L{max_row}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=False,
        showColumnStripes=False, showFirstColumn=False, showLastColumn=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A3"

    # 条件格式1：每张发票第一行蓝底（票号与上一行不同=新发票）
    from openpyxl.formatting.rule import FormulaRule
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.conditional_formatting.add(
        f"A3:L{max_row}",
        FormulaRule(formula=[f'$K3<>$K2'], fill=blue_fill),
    )

    # 条件格式2：续行的发票级信息（序号/日期/总额/票号/供应商）与上一行相同则隐藏
    hide_font = Font(color="FFFFFF")
    for col in ["A", "B", "J", "K", "L"]:
        rng = f"{col}3:{col}{max_row}"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'{col}3={col}2'], font=hide_font),
        )

    wb.save(output_path)


# ==================== 辅助函数 ====================
invoices_store = []


def _parse_amount(s):
    try:
        return float(str(s).replace(",", "").replace("￥", "").replace("¥", "").strip() or "0")
    except (ValueError, TypeError):
        return 0.0


def _sort_invoices(invoices, sort_by="date_desc"):
    if sort_by == "amount_desc":
        return sorted(invoices, key=lambda inv: _parse_amount(inv.get("total_amount", "0")), reverse=True)
    elif sort_by == "amount_asc":
        return sorted(invoices, key=lambda inv: _parse_amount(inv.get("total_amount", "0")))
    elif sort_by == "date_asc":
        return sorted(invoices, key=lambda inv: inv.get("date", "") or "")
    else:
        return sorted(invoices, key=lambda inv: inv.get("date", "") or "", reverse=True)


def _has_missing_fields(invoice):
    for item in invoice.get("items", []):
        amount = item.get("amount", "").strip()
        qty = item.get("qty", "").strip()
        price = item.get("price", "").strip()
        name = item.get("name", "").strip()
        spec = item.get("spec", "").strip()
        if amount and (not qty or not price):
            return True
        if not name or not spec:
            return True
        if name and not amount:
            return True
    return False


def _save_to_history(invoice, uploaded_by):
    conn = get_db()
    conn.execute(
        "INSERT INTO invoices (id, invoice_num, date, seller_name, total_amount, items, filename, uploaded_by, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
        (invoice["id"], invoice.get("invoice_num", ""), invoice.get("date", ""),
         invoice.get("seller_name", ""), invoice.get("total_amount", ""),
         json.dumps(invoice.get("items", []), ensure_ascii=False),
         invoice.get("filename", ""), uploaded_by, datetime.now().isoformat()),
    )
    conn.commit()
    conn.close()


def _check_duplicate(inv_num, username):
    for existing in invoices_store:
        if existing.get("invoice_num", "") == inv_num and existing.get("uploaded_by") == username:
            return existing.get("filename", "未知文件")
    conn = get_db()
    row = conn.execute("SELECT filename FROM invoices WHERE invoice_num = ? AND uploaded_by = ?",
                       (inv_num, username)).fetchone()
    conn.close()
    if row:
        return row["filename"]
    return None


def _get_user_invoices():
    """获取当前用户的发票列表"""
    username = session.get("username", "")
    return [inv for inv in invoices_store if inv.get("uploaded_by") == username]


# ==================== 认证路由 ====================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    data = request.json or request.form
    username = data.get("username", "").strip()
    password = data.get("password", "")

    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
    conn.close()

    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"error": "用户名或密码错误"}), 401

    session.permanent = True
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    return jsonify({"success": True, "redirect": "/"})


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")


# ==================== API Key 设置 ====================
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "GET":
        conn = get_db()
        user = conn.execute("SELECT deepseek_api_key FROM users WHERE id = ?",
                            (session["user_id"],)).fetchone()
        conn.close()
        has_key = bool(user and user["deepseek_api_key"])
        return render_template("settings.html", username=session.get("username", ""), has_key=has_key)

    data = request.json or request.form
    api_key = data.get("api_key", "").strip()
    if not api_key:
        return jsonify({"error": "请输入API Key"}), 400

    conn = get_db()
    conn.execute("UPDATE users SET deepseek_api_key = ? WHERE id = ?",
                 (api_key, session["user_id"]))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "redirect": "/"})


@app.route("/api/user/has_key")
@login_required
def has_api_key():
    conn = get_db()
    user = conn.execute("SELECT deepseek_api_key FROM users WHERE id = ?",
                        (session["user_id"],)).fetchone()
    conn.close()
    return jsonify({"has_key": bool(user and user["deepseek_api_key"])})


# ==================== 历史记录路由 ====================
@app.route("/api/history")
@login_required
def api_history():
    q = request.args.get("q", "").strip()
    limit = min(int(request.args.get("limit", 20)), 200)
    offset = int(request.args.get("offset", 0))
    username = session.get("username", "")

    conn = get_db()
    if q:
        like = f"%{q}%"
        rows = conn.execute(
            """SELECT id, invoice_num, date, seller_name, total_amount, filename, uploaded_by, created_at
               FROM invoices
               WHERE uploaded_by = ? AND (
                   date LIKE ? OR seller_name LIKE ? OR invoice_num LIKE ?
                   OR total_amount LIKE ?)
               ORDER BY created_at DESC LIMIT ? OFFSET ?""",
            (username, like, like, like, like, limit, offset),
        ).fetchall()
    else:
        rows = conn.execute(
            """SELECT id, invoice_num, date, seller_name, total_amount, filename, uploaded_by, created_at
               FROM invoices WHERE uploaded_by = ?
               ORDER BY created_at DESC LIMIT ? OFFSET ?""",
            (username, limit, offset),
        ).fetchall()
    conn.close()
    return jsonify({"history": [dict(r) for r in rows]})


@app.route("/api/history/<hid>")
@login_required
def api_history_detail(hid):
    conn = get_db()
    row = conn.execute("SELECT * FROM invoices WHERE id = ?", (hid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "未找到"}), 404
    inv = dict(row)
    inv["items"] = json.loads(inv["items"])
    return jsonify({"invoice": inv})


@app.route("/api/history/<hid>", methods=["DELETE"])
@login_required
def delete_history(hid):
    conn = get_db()
    conn.execute("DELETE FROM invoices WHERE id = ?", (hid,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ==================== 主页面路由 ====================
@app.route("/")
@login_required
def index():
    conn = get_db()
    user = conn.execute("SELECT deepseek_api_key FROM users WHERE id = ?",
                        (session["user_id"],)).fetchone()
    conn.close()
    if not user or not user["deepseek_api_key"]:
        return redirect("/settings")
    return render_template("index_v2.html",
                           username=session.get("username", ""))


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "未收到文件"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALL_EXTENSIONS:
        return jsonify({"error": f"不支持的格式: {ext}"}), 400

    safe_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(UPLOAD_DIR, safe_name)
    file.save(file_path)

    try:
        ocr_results = ocr_engine.recognize(file_path)

        user_api_key = get_user_api_key(session["user_id"])
        max_attempts = 3
        retry_count = 0
        for attempt in range(max_attempts):
            invoice = parse_with_deepseek(ocr_results, user_api_key)
            if not _has_missing_fields(invoice):
                break
            if attempt < max_attempts - 1:
                retry_count += 1
                print(f"  [重试 {retry_count}/{max_attempts-1}] {file.filename} 部分明细不完整，重新解析...")

        invoice["items"] = [item for item in invoice.get("items", [])
                            if item.get("amount", "").strip()
                            or item.get("qty", "").strip()
                            or item.get("price", "").strip()]

        # 检查重试后是否仍有空字段
        has_incomplete = _has_missing_fields(invoice)

        inv_num = invoice.get("invoice_num", "")
        if inv_num:
            dup_file = _check_duplicate(inv_num, session.get("username", ""))
            if dup_file:
                return jsonify({
                    "success": False, "duplicate": True,
                    "warning": f"重复发票！票号 {inv_num} 已存在（来自 {dup_file}），已跳过。",
                })

        invoice["id"] = uuid.uuid4().hex
        invoice["filename"] = file.filename
        invoice["uploaded_by"] = session.get("username", "")
        invoices_store.append(invoice)
        _save_to_history(invoice, session.get("username", ""))

        return jsonify({
            "success": True,
            "retry_count": retry_count,
            "has_incomplete": has_incomplete,
            "invoice": {
                "id": invoice["id"], "filename": invoice["filename"],
                "date": invoice["date"], "invoice_num": invoice["invoice_num"],
                "total_amount": invoice["total_amount"],
                "seller_name": invoice["seller_name"],
                "item_count": len(invoice["items"]),
                "items": invoice["items"],
            },
        })
    except Exception as e:
        return jsonify({"error": f"处理异常: {str(e)}"}), 200
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)


@app.route("/results")
@login_required
def results():
    user_invoices = _get_user_invoices()
    return jsonify({"invoices": [
        {"id": inv["id"], "filename": inv["filename"], "date": inv["date"],
         "invoice_num": inv["invoice_num"], "total_amount": inv["total_amount"],
         "seller_name": inv["seller_name"], "item_count": len(inv["items"]),
         "items": inv["items"]}
        for inv in user_invoices
    ]})


@app.route("/download")
@login_required
def download():
    user_invoices = _get_user_invoices()
    if not user_invoices:
        return jsonify({"error": "没有数据"}), 400
    filename = f"进项发票记录_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(UPLOAD_DIR, filename)
    write_excel(user_invoices, output_path)
    return send_file(output_path, as_attachment=True, download_name=filename)


@app.route("/download/history", methods=["POST"])
@login_required
def download_history():
    data = request.json or {}
    ids = data.get("ids", [])
    if not ids:
        return jsonify({"error": "未选择记录"}), 400

    username = session.get("username", "")
    conn = get_db()
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(
        f"SELECT * FROM invoices WHERE id IN ({placeholders}) AND uploaded_by = ?",
        (*ids, username),
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({"error": "未找到记录"}), 404

    invoices_by_id = {}
    for row in rows:
        inv = dict(row)
        inv["items"] = json.loads(inv["items"])
        invoices_by_id[inv["id"]] = inv

    invoices = [invoices_by_id[i] for i in ids if i in invoices_by_id]

    filename = f"进项发票记录_历史选中_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(UPLOAD_DIR, filename)
    write_excel(invoices, output_path)
    return send_file(output_path, as_attachment=True, download_name=filename)


@app.route("/clear")
@login_required
def clear():
    username = session.get("username", "")
    invoices_store[:] = [inv for inv in invoices_store if inv.get("uploaded_by") != username]
    return jsonify({"success": True})


@app.route("/delete/<invoice_id>")
@login_required
def delete_invoice(invoice_id):
    username = session.get("username", "")
    invoices_store[:] = [inv for inv in invoices_store
                         if not (inv["id"] == invoice_id and inv.get("uploaded_by") == username)]
    conn = get_db()
    conn.execute("DELETE FROM invoices WHERE id = ? AND uploaded_by = ?", (invoice_id, username))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


if __name__ == "__main__":
    init_db()
    print("=" * 55)
    print("  发票识别系统 V3 - OCR + DeepSeek LLM")
    print("  上海辉驰包装设备有限公司 财务部")
    print("=" * 55)
    print("\n  OCR识别 → DeepSeek理解表格 → 提取全部字段")
    print("  浏览器访问: http://localhost:8080\n")
    app.run(host="0.0.0.0", port=8080, debug=True, threaded=True)
